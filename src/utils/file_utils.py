import os
import json
from openpyxl import Workbook, load_workbook
from config.path import *

def write_agency_2_json(agency_name, other_names):
    agency_list = read_json(AGENCY_LIST_PATH)
    agency_inv_list = read_json(AGENCY_INV_LIST_PATH)
    if agency_name not in agency_list.keys():
        agency_list[agency_name]=[]
    if other_names != '' and other_names != None:
        for other in other_names.split('///'):
            if other != '' and other != None:
                if other not in agency_list[agency_name]:
                    agency_list[agency_name].append(other)
                if other not in agency_inv_list.keys():
                    agency_inv_list[other] = [agency_name]
                elif agency_name not in agency_inv_list[other]:
                    agency_inv_list[other].append(agency_name)
    save_json(agency_list, AGENCY_LIST_PATH)
    save_json(agency_inv_list, AGENCY_INV_LIST_PATH)

def makedir(path):
    if not os.path.exists(path):
        os.makedirs(path)
        print(f"文件夹 '{path}' 创建成功！")
    else:
        print(f"文件夹 '{path}' 已存在！")

def delete_file(path):
    if os.path.exists(path):
        try:
            os.remove(path)
        except OSError as e:
            print(f'删除文件{path}时出错：{e}')

def read_txt(path, sep):
    data = []
    with open(path, 'r', encoding='utf-8') as file:
        lines = file.readlines()
        if sep == '':
            for line in lines:
                line = line.strip()     # 去除首尾空白字符
                if not line:
                    continue
                data.append(line)
        elif sep != '':
            data = lines.split(sep)
            for i in enumerate(len(data)):
                data[i] = data[i].replace("\n", "").replace("\r", "")

    return data

def write_txt(path, textlist):
    with open(path, "w", encoding="utf-8") as f:
        for text in textlist:
            clean_text = text.replace("\n", "").replace("\r", "")
            f.write(clean_text + "\n")


# 读取 JSON 文件
def read_json(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)
    return data

# 保存数据到 JSON 文件
def save_json(data, file_path):
    with open(file_path, 'w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)


def create_excel(cols, data, title="file", path=None):
    # 创建一个新的工作簿
    wb = Workbook()

    # 获取默认的工作表（Sheet）
    ws = wb.active

    # 设置工作表的标题
    ws.title = title

    # 写入数据到工作表
    # 写入表头
    for i, col in enumerate(cols):
        ws[chr(65+i) + '1'] = col
    
    # 从第二行开始写入数据
    for row in data:
        ws.append(row)

    # 保存工作簿到文件
    #wb.save(f"/disk2/zyy/grad_pro/shared/task_output_excel/{task_id}_.xlsx")
    if path:
        wb.save(path)
        print(f"Excel 文件已创建并保存为 {path}")
    else:
        print(f'待转化为excel的文件不存在!')


def read_excel(path):
    # 加载 Excel 文件
    wb = load_workbook(path)

    # 获取活动工作表（也可以通过 wb['SheetName'] 获取特定工作表）
    ws = wb.active

    # 读取表头（第一行）
    cols = [cell.value for cell in ws[1]]

    # 读取数据（从第二行开始）
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(list(row))

    return cols, data   

def excel_2_txt(file, task_id, start_row, col):
    wb = load_workbook(file)
    ws = wb.active
    datalist = [[cell.value for cell in row] for row in ws.iter_rows()][start_row - 1:]
    textlist = [data[col-1] for data in datalist]
    write_txt(f'{manual_input_folder}{task_id}_.txt', textlist)
    
def taskdata_output_2_excel(task_id, datalist):
  rows = []
  for data in datalist:
    rows.append((data.data_id, data.platform, data.time, data.text, data.output))
  cols = ['评论编号', '平台', '发布时间', '评论内容', '关系输出']
  create_excel(cols, rows, title=f'任务{task_id}数据', path=f'{task_output_excel_folder}{task_id}_.xlsx')

def taskdata_output_2_excel_manual(task_id, trans_rels_list):
  datalist = read_json(f"{manual_filtered_model_output_folder}{task_id}_.rel")
  key_2_output = {}
  rows = []
  for i, data in enumerate(datalist):
    key_raw = "_".join(data["doc_key"].split("_")[:-1])
    if key_raw not in key_2_output.keys():
        key_2_output[key_raw] = [data["text"], []]
    if trans_rels_list[i] not in key_2_output[key_raw][1]:
        key_2_output[key_raw][1].append(trans_rels_list[i])

  for key, data in key_2_output.items():
    rows.append((key, data[0], str(data[1])))
  cols = ['评论编号', '评论内容', '关系输出']
  create_excel(cols, rows, title=f'任务{task_id}数据', path=f'{manual_task_output_excel_folder}{task_id}_.xlsx')

